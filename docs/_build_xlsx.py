"""
Build the KU-branded SPMD Minor → Fall 2027 ADEK Roadmap Excel workbook.

Reads  ../_data.json
Writes docs/SPMD_Fall2027_Roadmap.xlsx

Sheets:
  1. Cover
  2. Roadmap      — Quarter × Track × Gate matrix
  3. Courses      — Full course catalogue
  4. Regulators   — 6 regulators × 4 concentration courses
  5. PLO_CLO      — Course-level PLO × CLO matrix
  6. Data_Issues  — Known data issues
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
DATA_PATH = HERE.parent / "_data.json"
OUT_PATH = HERE / "SPMD_Fall2027_Roadmap.xlsx"

GEN_DATE = "2026-05-18"
FOOTER_TEXT = "BSCND Sports Medicine Roadmap · Fall 2027 · Generated 2026-05-18 · ADEK-aligned"

# ---------------------------------------------------------------------------
# Brand palette
# ---------------------------------------------------------------------------
NAVY = "003D7A"
GOLD = "C8A04A"
GOLD_LIGHT = "E8D08A"
WHITE = "FFFFFF"
GREY_BG = "F3F4F6"
AMBER = "D97706"
AMBER_LIGHT = "FFF7ED"
GREEN = "1B8A4E"
INK = "1F2937"

navy_fill = PatternFill("solid", fgColor=NAVY)
gold_fill = PatternFill("solid", fgColor=GOLD)
amber_fill = PatternFill("solid", fgColor=AMBER)
amber_light_fill = PatternFill("solid", fgColor=AMBER_LIGHT)
grey_fill = PatternFill("solid", fgColor=GREY_BG)
green_fill = PatternFill("solid", fgColor=GREEN)

FONT_BASE = "Calibri"

title_font = Font(name=FONT_BASE, size=18, bold=True, color=WHITE)
header_font = Font(name=FONT_BASE, size=11, bold=True, color=WHITE)
subheader_font = Font(name=FONT_BASE, size=12, bold=True, color=WHITE)
body_font = Font(name=FONT_BASE, size=10)
body_bold = Font(name=FONT_BASE, size=10, bold=True)
navy_font = Font(name=FONT_BASE, size=10, bold=True, color=NAVY)
amber_font = Font(name=FONT_BASE, size=10, bold=True, color=AMBER)
small_font = Font(name=FONT_BASE, size=9)
tiny_font = Font(name=FONT_BASE, size=8)

thin_side = Side(style="thin", color="BFBFBF")
gold_side = Side(style="medium", color=GOLD)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
gold_bottom = Border(bottom=gold_side)
gold_left = Border(left=Side(style="medium", color=GOLD), right=thin_side,
                   top=thin_side, bottom=thin_side)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def color_tab(ws, hex_color: str = NAVY) -> None:
    ws.sheet_properties.tabColor = hex_color


def set_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_band(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = navy_fill
    c.font = title_font
    c.alignment = center
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = gold_bottom
    ws.row_dimensions[row].height = 36


def gold_band(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = gold_fill
    c.font = subheader_font
    c.alignment = left_mid
    ws.row_dimensions[row].height = 22


def header_row(ws, row: int, headers: list, col_start: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.fill = navy_fill
        c.font = header_font
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[row].height = 22


def body_cell(ws, row: int, col: int, value: str,
              fill=None, bold: bool = False,
              align: Alignment = None,
              font_override: Font = None,
              is_first_col: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_override or (body_bold if bold else body_font)
    c.alignment = align or left_wrap
    c.border = gold_left if is_first_col else border_all
    if fill:
        c.fill = fill


def alt_fill(row_index: int) -> PatternFill | None:
    return grey_fill if row_index % 2 == 0 else None


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def build_cover(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Cover")
    color_tab(ws)
    set_widths(ws, [22, 55, 18, 18])

    title_band(ws, 1, 1, 4, "KHALIFA UNIVERSITY — SPMD Minor → Fall 2027 ADEK Roadmap")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sub = ws.cell(row=2, column=1,
                  value="College of Medicine and Health Sciences  ·  BSc Clinical Nutrition & Dietetics  ·  ADEK Submission")
    sub.fill = navy_fill
    sub.font = Font(name=FONT_BASE, size=12, italic=True, color=WHITE)
    sub.alignment = center
    ws.row_dimensions[2].height = 24

    prog = data["program"]
    meta_pairs = [
        ("Programme", prog["full_name"]),
        ("Institution", prog["institution"]),
        ("College", prog["college"]),
        ("Target Launch", prog["launch"]),
        ("QF Emirates Level", str(data["qfemirates_level"])),
        ("Baseline Credits", str(prog["total_credits_baseline"])),
        ("Credits with Concentration", str(prog["total_credits_with_concentration"])),
        ("Prepared by", "Dr. Carlo Raj"),
        ("Date", GEN_DATE),
        ("Audience", "ADEK Reviewers"),
    ]

    gold_band(ws, 4, 1, 4, "Programme Metadata")
    for i, (k, v) in enumerate(meta_pairs, start=1):
        r = 4 + i
        kc = ws.cell(row=r, column=1, value=k)
        kc.fill = grey_fill
        kc.font = body_bold
        kc.alignment = left_mid
        kc.border = border_all
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        vc = ws.cell(row=r, column=2, value=v)
        vc.font = body_font
        vc.alignment = left_mid
        vc.border = border_all
        for col in (3, 4):
            ws.cell(row=r, column=col).border = border_all

    legend_row = 4 + len(meta_pairs) + 2
    gold_band(ws, legend_row, 1, 4, "Legend")
    header_row(ws, legend_row + 1, ["Symbol / Colour", "Meaning", "Context", "Notes"])
    legend_items = [
        ("Navy fill", "Table header / emphasis", "All tables", ""),
        ("Gold fill", "Section header / sub-band", "All sections", ""),
        ("Amber fill", "PROPOSED — inferred data", "SPMD 305–308", "Requires faculty validation"),
        ("Green fill", "Status: Approved / confirmed", "Status cells", ""),
        ("● (bullet)", "CLO maps to this PLO", "PLO×CLO matrix", ""),
        ("—", "No alignment", "Regulator matrix", ""),
    ]
    for j, (sym, meaning, ctx, notes) in enumerate(legend_items):
        r = legend_row + 2 + j
        fills = [None, None, None, None]
        if "Navy" in sym:
            fills[0] = navy_fill
        elif "Gold" in sym:
            fills[0] = gold_fill
        elif "Amber" in sym:
            fills[0] = amber_fill
        elif "Green" in sym:
            fills[0] = green_fill
        for col_idx, (val, fill) in enumerate(zip([sym, meaning, ctx, notes], fills), start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = body_font
            c.alignment = left_mid
            c.border = border_all
            if fill:
                c.fill = fill
                c.font = Font(name=FONT_BASE, size=10, color=WHITE)

    nav_row = legend_row + len(legend_items) + 4
    gold_band(ws, nav_row, 1, 4, "Navigation — Sheet Index")
    nav_items = [
        ("Roadmap", "Quarter × Track × Gate implementation matrix"),
        ("Courses", "Full course catalogue with PLOs, CLOs, and status"),
        ("Regulators", "Six UAE regulators × four concentration courses"),
        ("PLO_CLO", "Course-level PLO × CLO mapping matrices"),
        ("Data_Issues", "Known data issues requiring decisions"),
    ]
    for k, (sheet, desc) in enumerate(nav_items, start=1):
        r = nav_row + k
        lc = ws.cell(row=r, column=1, value=sheet)
        lc.hyperlink = f"#'{sheet}'!A1"
        lc.font = Font(name=FONT_BASE, size=11, bold=True, color=NAVY, underline="single")
        lc.alignment = left_mid
        lc.border = border_all
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        dc = ws.cell(row=r, column=2, value=desc)
        dc.font = body_font
        dc.alignment = left_mid
        dc.border = border_all
        for col in (3, 4):
            ws.cell(row=r, column=col).border = border_all

    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 2: Roadmap
# ---------------------------------------------------------------------------

def build_roadmap(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Roadmap")
    color_tab(ws, GOLD)
    set_widths(ws, [10, 20, 35, 35, 28, 30])

    title_band(ws, 1, 1, 6, "Quarter × Track × Gate Implementation Roadmap — Fall 2027")

    # Merged sub-headers: Quarter | Label | Minor | Concentration | Gate | Owner
    # Row 2: merged header groups
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    for col, label in [(1, "Quarter"), (2, "Phase Label / Dates"), (3, "Track Deliverables"), (5, "Governance Gate"), (6, "Owner / Decision Maker")]:
        c = ws.cell(row=2, column=col, value=label)
        c.fill = navy_fill
        c.font = header_font
        c.alignment = center
        c.border = border_all
    ws.cell(row=2, column=4).fill = navy_fill
    ws.cell(row=2, column=4).border = border_all

    headers = ["Quarter", "Label / Dates", "Minor (SPMD 301–304) Deliverable",
               "Concentration (SPMD 305–308) Deliverable", "Governance Gate", "Owner / Decision Maker"]
    header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 28

    phases = data["phases"]
    for i, phase in enumerate(phases, start=1):
        r = 3 + i
        vals = [
            phase["q"],
            f"{phase['label']}\n{phase['date_range']}",
            phase["deliverables"]["minor"],
            phase["deliverables"]["concentration"],
            phase["gate"],
            f"{phase['owner']}\nDecision-maker: {phase['decision_maker']}",
        ]
        fill = amber_light_fill if phase["q"] == "Q3 2027" else (grey_fill if i % 2 == 0 else None)
        fnt = amber_font if phase["q"] == "Q3 2027" else body_font
        for j, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = fnt if j > 1 else (Font(name=FONT_BASE, size=10, bold=True, color=NAVY)
                                         if phase["q"] == "Q3 2027" else body_bold)
            c.alignment = center if j == 1 else left_wrap
            c.border = gold_left if j == 1 else border_all
            if fill:
                c.fill = fill
        ws.row_dimensions[r].height = 60

    # Key actions block
    ka_row = 3 + len(phases) + 2
    gold_band(ws, ka_row, 1, 6, "Key Actions by Quarter")
    r = ka_row + 1
    header_row(ws, r, ["Quarter", "Key Action", "", "", "", ""], col_start=1)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    for phase in phases:
        for action in phase.get("key_actions", []):
            c1 = ws.cell(row=r, column=1, value=phase["q"])
            c1.font = navy_font
            c1.alignment = center
            c1.border = gold_left
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c2 = ws.cell(row=r, column=2, value=f"• {action}")
            c2.font = small_font
            c2.alignment = left_wrap
            c2.border = border_all
            for col in range(3, 7):
                ws.cell(row=r, column=col).border = border_all
            r += 1

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 3: Courses
# ---------------------------------------------------------------------------

def build_courses(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Courses")
    color_tab(ws)
    set_widths(ws, [10, 32, 8, 10, 14, 18, 50, 55, 14, 16])

    title_band(ws, 1, 1, 10, "SPMD Course Catalogue — SPMD 301–308")

    header_row(ws, 2, ["Code", "Title", "Credits", "Level", "Semester",
                       "Track / Status", "Description", "CLOs",
                       "PLOs", "KU GELOs"])
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    courses = data["courses"]
    for i, code in enumerate(sorted(courses.keys()), start=1):
        c = courses[code]
        r = 2 + i
        is_inferred = c.get("_inferred", False)
        status_text = c.get("status", "")
        if is_inferred:
            status_text = "PROPOSED — pending faculty validation"
        track_status = f"{c.get('track', '').title()}\n{status_text}"

        clos_text = "\n".join(c.get("clos", []))
        plos_text = ", ".join(sorted(c.get("plos", [])))
        gelos_text = ", ".join(c.get("ku_gelo", []))

        fill = amber_light_fill if is_inferred else (grey_fill if i % 2 == 0 else None)

        vals = [
            code,
            c["title"],
            str(c["credits"]),
            str(c.get("level", "")),
            c.get("semester", ""),
            track_status,
            c.get("description", ""),
            clos_text,
            plos_text,
            gelos_text,
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = center_top if j in (1, 3, 4, 9, 10) else left_wrap
            cell.border = gold_left if j == 1 else border_all
            if fill:
                cell.fill = fill
            if is_inferred and j == 6:
                cell.font = amber_font
            elif j == 1:
                cell.font = navy_font
            else:
                cell.font = small_font if j in (7, 8) else body_font

        ws.row_dimensions[r].height = max(80, 13 * len(c.get("clos", [])))


# ---------------------------------------------------------------------------
# Sheet 4: Regulators
# ---------------------------------------------------------------------------

def build_regulators(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Regulators")
    color_tab(ws, GOLD)
    set_widths(ws, [12, 26, 18, 12, 35, 35, 35, 35])

    conc_codes = sorted(["SPMD305", "SPMD306", "SPMD307", "SPMD308"])
    headers = ["Regulator", "Full Name", "Tier / Level", "Scope"] + conc_codes

    title_band(ws, 1, 1, len(headers), "UAE Regulator Alignment Matrix — Concentration Courses (SPMD 305–308)")
    header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    regulators = data["regulators"]
    for i, reg_key in enumerate(sorted(regulators.keys()), start=1):
        reg = regulators[reg_key]
        alignment = reg.get("alignment", {})
        r = 2 + i
        fill = grey_fill if i % 2 == 0 else None

        base_vals = [
            reg_key,
            reg.get("full_name", ""),
            reg.get("tier", ""),
            reg.get("scope", ""),
        ]
        for j, val in enumerate(base_vals, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.alignment = center if j == 1 else left_wrap
            c.border = gold_left if j == 1 else border_all
            c.font = navy_font if j == 1 else (tiny_font if j == 4 else small_font)
            if fill:
                c.fill = fill

        for j, code in enumerate(conc_codes):
            col = 5 + j
            align_text = alignment.get(code, "")
            c = ws.cell(row=r, column=col, value=align_text or "—")
            c.alignment = left_wrap
            c.border = border_all
            c.font = tiny_font
            if not align_text:
                c.fill = grey_fill
                c.font = Font(name=FONT_BASE, size=9, color="9CA3AF")
            elif fill:
                c.fill = fill

        ws.row_dimensions[r].height = 55

    # Regulator detail section
    detail_start = 2 + len(regulators) + 2
    gold_band(ws, detail_start, 1, len(headers), "Regulator Details")
    header_row(ws, detail_start + 1, ["Regulator", "Full Name (AR)", "Level", "URL",
                                       "Requirements (REPs UAE only)", "", "", ""])
    ws.merge_cells(start_row=detail_start + 1, start_column=5,
                   end_row=detail_start + 1, end_column=len(headers))
    r = detail_start + 2
    for reg_key in sorted(regulators.keys()):
        reg = regulators[reg_key]
        reqs = reg.get("requirements", [])
        reqs_text = "\n".join(f"• {req}" for req in reqs) if reqs else ""
        vals = [reg_key, reg.get("name_ar", ""), reg.get("level", ""), reg.get("url", ""), reqs_text]
        for j, val in enumerate(vals, start=1):
            col_end = len(headers) if j == 5 else j
            if j == 5:
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=len(headers))
            c = ws.cell(row=r, column=j, value=val)
            c.font = navy_font if j == 1 else tiny_font
            c.alignment = center if j == 1 else left_wrap
            c.border = gold_left if j == 1 else border_all
        ws.row_dimensions[r].height = max(40, 14 * len(reqs))
        r += 1


# ---------------------------------------------------------------------------
# Sheet 5: PLO_CLO
# ---------------------------------------------------------------------------

def build_plo_clo(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("PLO_CLO")
    color_tab(ws)
    set_widths(ws, [10, 60, 10, 10, 10, 10, 10])

    plos = data["bscnd_plos"]
    plo_ids = [p["id"] for p in plos]

    title_band(ws, 1, 1, 2 + len(plo_ids), "PLO × CLO Matrix — SPMD 301–308")

    # PLO header sub-row
    gold_band(ws, 2, 1, 2 + len(plo_ids), "Programme Learning Outcomes (PLOs)")
    header_row(ws, 3, ["Course / CLO", "CLO Text"] + plo_ids)
    ws.freeze_panes = "A4"

    r = 4
    courses = data["courses"]
    for code in sorted(courses.keys()):
        c = courses[code]
        is_inferred = c.get("_inferred", False)
        course_plos = set(c.get("plos", []))
        clos = c.get("clos", [])

        # Course header row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(plo_ids))
        hc = ws.cell(row=r, column=1,
                     value=f"{code} — {c['title']}" + (" [PROPOSED — pending faculty validation]" if is_inferred else ""))
        hc.fill = amber_light_fill if is_inferred else navy_fill
        hc.font = Font(name=FONT_BASE, size=11, bold=True,
                       color=AMBER if is_inferred else WHITE)
        hc.alignment = left_mid
        for col in range(1, 3 + len(plo_ids)):
            ws.cell(row=r, column=col).border = border_all
        ws.row_dimensions[r].height = 22
        r += 1

        for clo_text in clos:
            # Extract CLO number (e.g. "CLO1:")
            clo_code = clo_text.split(":")[0].strip() if ":" in clo_text else ""
            body_text = clo_text[len(clo_code) + 1:].strip() if clo_code else clo_text

            c1 = ws.cell(row=r, column=1, value=clo_code)
            c1.font = body_bold
            c1.alignment = center
            c1.border = gold_left

            c2 = ws.cell(row=r, column=2, value=body_text)
            c2.font = small_font
            c2.alignment = left_wrap
            c2.border = border_all

            for j, plo_id in enumerate(plo_ids):
                col = 3 + j
                cell = ws.cell(row=r, column=col)
                cell.alignment = center
                cell.border = border_all
                if plo_id in course_plos:
                    cell.fill = navy_fill
                    cell.value = "●"
                    cell.font = Font(name=FONT_BASE, size=11, bold=True, color=WHITE)
                else:
                    cell.fill = grey_fill
                    cell.value = ""
                    cell.font = small_font
            ws.row_dimensions[r].height = 30
            r += 1

        # Spacer
        for col in range(1, 3 + len(plo_ids)):
            ws.cell(row=r, column=col).border = border_all
        r += 1


# ---------------------------------------------------------------------------
# Sheet 6: Data_Issues
# ---------------------------------------------------------------------------

def build_data_issues(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Data_Issues")
    color_tab(ws, AMBER)
    set_widths(ws, [8, 90])

    title_band(ws, 1, 1, 2, "Known Data Issues — Decision Required Before UGCC Submission")

    header_row(ws, 2, ["#", "Issue — Decision needed"])
    ws.freeze_panes = "A3"

    issues = data.get("known_data_issues", [])
    for i, issue in enumerate(issues, start=1):
        r = 2 + i
        c1 = ws.cell(row=r, column=1, value=str(i))
        c1.font = body_bold
        c1.alignment = center
        c1.border = gold_left
        c1.fill = amber_light_fill

        c2 = ws.cell(row=r, column=2, value=f"Decision needed: {issue}")
        c2.font = small_font
        c2.alignment = left_wrap
        c2.border = border_all
        ws.row_dimensions[r].height = max(30, 14 * ((len(issue) // 80) + 1))

    if not issues:
        ws.cell(row=3, column=1, value="(No data issues recorded)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    with open(DATA_PATH, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_cover(wb, data)
    build_roadmap(wb, data)
    build_courses(wb, data)
    build_regulators(wb, data)
    build_plo_clo(wb, data)
    build_data_issues(wb, data)

    # Set Cover as active on open
    wb.active = wb.index(wb["Cover"])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    size = OUT_PATH.stat().st_size
    print(f"[OK] Wrote: {OUT_PATH}  ({size:,} bytes)")


if __name__ == "__main__":
    main()
